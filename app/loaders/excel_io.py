"""Robust Excel file I/O with merged cell handling."""
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import logging

from app.utils.fa_normalize import normalize_fa

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def load_workbook(filepath: str):
    """
    Load an Excel workbook with error handling.
    
    Args:
        filepath: Path to Excel file
        
    Returns:
        openpyxl Workbook object
        
    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If file cannot be loaded
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        logger.info(f"Loaded workbook: {filepath}")
        return wb
    except FileNotFoundError:
        logger.error(f"File not found: {filepath}")
        raise
    except Exception as e:
        logger.error(f"Error loading {filepath}: {e}")
        raise


def get_sheet_grid(sheet, normalize: bool = True) -> List[List[Any]]:
    """
    Convert an Excel sheet to a 2D grid (list of lists).
    Handles merged cells by forward-filling values.
    
    Args:
        sheet: openpyxl worksheet object
        normalize: If True, normalize Persian text in cells
        
    Returns:
        2D list representing the sheet grid
    """
    # Get all merged cell ranges
    merged_ranges = {}
    for merged_range in sheet.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        
        # Get the value from the top-left cell of merged range
        top_left_cell = sheet.cell(min_row, min_col)
        value = top_left_cell.value
        
        # Store for all cells in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_ranges[(row, col)] = value
    
    # Build the grid
    grid = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        grid_row = []
        for col_idx, cell in enumerate(row, start=1):
            # Check if this cell is part of a merged range
            if (row_idx, col_idx) in merged_ranges:
                value = merged_ranges[(row_idx, col_idx)]
            else:
                value = cell.value
            
            # Normalize if requested
            if normalize and value is not None:
                value = normalize_cell(value)
            
            grid_row.append(value)
        grid.append(grid_row)
    
    return grid


def normalize_cell(value: Any) -> Any:
    """
    Normalize a single cell value.
    
    Args:
        value: Cell value (can be string, number, None, etc.)
        
    Returns:
        Normalized value
    """
    if value is None:
        return None
    
    if isinstance(value, str):
        normalized = normalize_fa(value)
        return normalized if normalized else None
    
    return value


def find_sheet_by_keyword(wb, keywords: List[str]) -> Optional[Any]:
    """
    Find a sheet in workbook that contains any of the keywords in its name.
    
    Args:
        wb: openpyxl workbook
        keywords: List of keywords to search for (case-insensitive)
        
    Returns:
        First matching sheet or None
    """
    for sheet in wb.worksheets:
        sheet_name = normalize_fa(sheet.title.lower())
        for keyword in keywords:
            keyword_norm = normalize_fa(keyword.lower())
            if keyword_norm in sheet_name:
                logger.info(f"Found sheet '{sheet.title}' matching keyword '{keyword}'")
                return sheet
    
    logger.warning(f"No sheet found matching keywords: {keywords}")
    return None


def get_first_sheet(wb):
    """Get the first worksheet from workbook."""
    return wb.worksheets[0]


def grid_to_dicts(grid: List[List[Any]], header_row_idx: int = 0) -> List[Dict[str, Any]]:
    """
    Convert a grid to a list of dictionaries using a header row.
    
    Args:
        grid: 2D list representing the sheet
        header_row_idx: Index of the header row (0-based)
        
    Returns:
        List of dictionaries, one per row
    """
    if not grid or len(grid) <= header_row_idx:
        return []
    
    headers = grid[header_row_idx]
    
    # Clean up headers
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers)]
    
    result = []
    for row in grid[header_row_idx + 1:]:
        # Skip empty rows
        if not any(row):
            continue
        
        row_dict = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = value
        
        result.append(row_dict)
    
    return result


def is_row_empty(row: List[Any]) -> bool:
    """Check if a row is empty (all None or empty strings)."""
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def find_header_row(grid: List[List[Any]], keywords: List[str], max_search_rows: int = 20) -> int:
    """
    Find the row index that contains header keywords.
    
    Args:
        grid: 2D grid
        keywords: List of keywords that should appear in header
        max_search_rows: Maximum number of rows to search
        
    Returns:
        Index of header row (0-based), or 0 if not found
    """
    search_limit = min(max_search_rows, len(grid))
    
    for idx in range(search_limit):
        row = grid[idx]
        row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
        
        # Check if any keyword is in this row
        matches = sum(1 for kw in keywords if normalize_fa(kw.lower()) in normalize_fa(row_text))
        
        if matches >= len(keywords) * 0.5:  # At least 50% of keywords match
            logger.info(f"Found header row at index {idx}")
            return idx
    
    logger.warning("Header row not found, assuming row 0")
    return 0
