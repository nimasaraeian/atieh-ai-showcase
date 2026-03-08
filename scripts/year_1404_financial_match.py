# -*- coding: utf-8 -*-
"""
خواندن کامل دیتای payments_<YEAR>_full.xlsx از ستون نام بیمار و تطبیق با فایل
نوبت‌دهی همان سال؛ گزارش تعداد نفراتی که وضعیت مالی دارند.

اجرا: python scripts/year_1404_financial_match.py [YEAR]
      python scripts/year_1404_financial_match.py 1403
      python scripts/year_1404_financial_match.py 1404
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).parent.parent))

import re
import pandas as pd
from openpyxl import load_workbook

# ─── مسیرها ─────────────────────────────────────────────────────────────────
REPO = Path(__file__).parent.parent
DEFAULT_YEAR = 1404


def _norm_header(s) -> str:
    """نرمال کردن هدر ستون برای تطبیق با نام‌های شناخته‌شده."""
    if s is None:
        return ""
    t = str(s).strip().replace("|", " ").replace("ي", "ی").replace("ك", "ک")
    t = " ".join(t.split())
    return t


def _norm_name_for_match(name: str) -> str:
    """برای تطبیق نام: نرمال کردن و حذف شماره پرونده در پرانتز."""
    if not name or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).strip()
    if s in ("", "nan", "None"):
        return ""
    # حذف (عدد) یا (تشکیل پرونده شده) از انتها
    s = re.sub(r"\s*\(\d+\)\s*$", "", s)
    s = re.sub(r"\s*\(تشکیل پرونده شده\)\s*$", "", s, flags=re.IGNORECASE)
    s = s.replace("ي", "ی").replace("ك", "ک").strip()
    s = " ".join(s.split())
    return s


def read_payments_patient_names(excel_path: Path) -> tuple[list[str], set[str]]:
    """
    خواندن تمام ردیف‌های payments و استخراج ستون نام بیمار (با read_only برای سرعت).
    برمی‌گرداند: (لیست همه نام‌ها به ترتیب ردیف، مجموعه نام‌های یونیک نرمال‌شده).
    """
    sheet_name = "MSExcel"
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    # اولین ردیف = هدر
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row:
        wb.close()
        raise ValueError("هدر خالی است.")
    col_index = {_norm_header(c): i for i, c in enumerate(header_row)}
    candidates = ["نام بيمار", "نام بیمار", "نام بیمار(تشکیل پرونده شده)"]
    col_idx = None
    for cand in candidates:
        n = _norm_header(cand)
        if n in col_index:
            col_idx = col_index[n]
            break
    if col_idx is None:
        for cand in candidates:
            n = _norm_header(cand)
            for k, v in col_index.items():
                if n in k or k in n:
                    col_idx = v
                    break
            if col_idx is not None:
                break
    if col_idx is None:
        wb.close()
        raise ValueError(f"ستون نام بیمار در فایل پرداخت پیدا نشد. ستون‌ها: {list(header_row)[:10]}...")
    all_names = []
    unique_norm = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_idx < len(row):
            val = row[col_idx]
        else:
            val = None
        s = _norm_name_for_match(val)
        if s:
            all_names.append(s)
            unique_norm.add(s)
    wb.close()
    return all_names, unique_norm


def read_history_patient_names(history_dir: Path) -> tuple[list[str], set[str]]:
    """
    پیدا کردن فایل اکسلی نوبت‌دهی و خواندن ستون نام بیمار.
    برمی‌گرداند: (لیست نام‌ها، مجموعه یونیک نرمال‌شده).
    """
    if not history_dir.exists():
        return [], set()
    xlsx_files = list(history_dir.glob("*.xlsx"))
    if not xlsx_files:
        return [], set()
    excel_path = sorted(xlsx_files)[0]
    # اولین شیت
    df = pd.read_excel(excel_path, sheet_name=0, engine="openpyxl", dtype=str)
    col_index = {_norm_header(c): c for c in df.columns}
    candidates = [
        "نام بیمار(تشکیل پرونده شده)",
        "نام بيمار(تشكيل پرونده شده)",
        "نام و نام خانوادگی",
        "نام بیمار",
    ]
    col = None
    for cand in candidates:
        n = _norm_header(cand)
        if n in col_index:
            col = col_index[n]
            break
    if col is None:
        for cand in candidates:
            n = _norm_header(cand)
            for k, v in col_index.items():
                if len(k) >= 10 and (n in k or k in n):
                    col = v
                    break
            if col is not None:
                break
    if col is None:
        return list(df.iloc[:, 0]), set()  # fallback first column
    all_names = []
    unique_norm = set()
    for val in df[col]:
        s = _norm_name_for_match(val)
        if s:
            all_names.append(s)
            unique_norm.add(s)
    return all_names, unique_norm


PAYMENTS_DIR = REPO / "data" / "inputs" / "payments"
HISTORY_BASE = REPO / "data" / "inputs" / "history"


def main():
    argv = sys.argv[1:]
    if argv and argv[0].lower() in ("all", "--all", "همه"):
        # حالت همه سال‌ها: جمع نام‌های منطبق (مالی + نوبتی) در هر سال، سپس یونیک کل
        payments_dir = PAYMENTS_DIR
        if not payments_dir.exists():
            print("پوشه پرداخت‌ها یافت نشد.")
            return
        years_from_files = set()
        for p in payments_dir.glob("payments_*_full.xlsx"):
            try:
                y = int(p.stem.replace("payments_", "").replace("_full", ""))
                years_from_files.add(y)
            except ValueError:
                pass
        years = sorted(years_from_files)
        all_matched_names: set[str] = set()
        for year in years:
            payments_file = payments_dir / f"payments_{year}_full.xlsx"
            history_dir = HISTORY_BASE / str(year)
            if not payments_file.exists():
                continue
            pay_names, pay_unique = read_payments_patient_names(payments_file)
            hist_names, hist_unique = read_history_patient_names(history_dir)
            overlap = pay_unique & hist_unique if hist_unique else set()
            all_matched_names |= overlap
            print(f"سال {year}: منطبق (مالی+نوبتی) = {len(overlap)} نفر")
        print()
        print("--- نتیجه کل ---")
        print(f"مجموع بیماران یونیک که در حداقل یک سال هم وضعیت مالی دارند هم در نوبت‌دهی منطبق‌اند: {len(all_matched_names)} نفر")
        return

    year = int(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_YEAR
    payments_file = REPO / "data" / "inputs" / "payments" / f"payments_{year}_full.xlsx"
    history_dir = REPO / "data" / "inputs" / "history" / str(year)

    print(f"=== تطبیق وضعیت مالی سال {year} ===\n")
    if not payments_file.exists():
        print(f"فایل پرداخت یافت نشد: {payments_file}")
        return
    print(f"خواندن پرداخت‌ها: {payments_file.name}")
    pay_names, pay_unique = read_payments_patient_names(payments_file)
    print(f"  تعداد ردیف (با نام غیرخالی): {len(pay_names)}")
    print(f"  تعداد نام یونیک (نرمال‌شده): {len(pay_unique)}")
    print()
    hist_names, hist_unique = read_history_patient_names(history_dir)
    overlap = pay_unique & hist_unique if hist_unique else set()
    if hist_unique:
        hist_file = sorted(history_dir.glob("*.xlsx"))[0]
        print(f"خواندن نوبت‌دهی {year}: {hist_file.name}")
        print(f"  تعداد ردیف با نام: {len(hist_names)}")
        print(f"  تعداد نام یونیک: {len(hist_unique)}")
        only_pay = pay_unique - hist_unique
        only_hist = hist_unique - pay_unique
        print()
        print("تطبیق نام (بر اساس نام نرمال‌شده):")
        print(f"  هم در پرداخت هم در نوبت‌دهی: {len(overlap)} نفر")
        print(f"  فقط در پرداخت (وضعیت مالی بدون نوبت در این فایل): {len(only_pay)} نفر")
        print(f"  فقط در نوبت‌دهی (بدون رکورد پرداخت در فایل {year}): {len(only_hist)} نفر")
    else:
        print(f"فایل نوبت‌دهی {year} یافت نشد یا ستون نام بیمار پیدا نشد.")
    print()
    print("--- نتیجه ---")
    print(f"مجموعاً در سال {year} تعداد {len(pay_unique)} نفر وضعیت مالی (رکورد پرداخت) دارند.")
    if hist_unique and overlap:
        print(f"از این میان {len(overlap)} نفر در فایل نوبت‌دهی {year} هم حضور دارند.")


if __name__ == "__main__":
    main()
