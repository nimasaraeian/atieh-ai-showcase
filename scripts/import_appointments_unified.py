# -*- coding: utf-8 -*-
"""
Unified import of all yearly appointment Excel files into appointments_unified_staging.
Files: نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_*.xlsx (and 1403: حضور پیدا کردند).
Idempotent: per-file replace (delete by source_file then insert).
"""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path
from datetime import datetime

REPO = Path(__file__).resolve().parent.parent
SQL_DIR = REPO / "sql" / "identity_resolution"
HISTORY_BASE = REPO / "data" / "inputs" / "history"
BATCH_SIZE = 2000

# Same file list as import_history_batch / build_appointment_recordno_bridge
APPOINTMENT_FILE_PATTERNS = [
    ("1395/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1395.xlsx", 1395),
    ("1396/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1396.xlsx", 1396),
    ("1398/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1398.xlsx", 1398),
    ("1399/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1399.xlsx", 1399),
    ("1400/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1400.xlsx", 1400),
    ("1401/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1401.xlsx", 1401),
    ("1402/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1402.xlsx", 1402),
    ("1403/نوبت_دهی_بیمارانی_که_حضور_پیدا_کردند_1403.xlsx", 1403),
    ("1404/نوبت_دهی_بیمارانی_که_حضور_پیدا_میکنند_1404.xlsx", 1404),
]

NAME_HEADERS = [
    "نام بیمار", "نام بيمار", "نام و نام خانوادگی", "نام بیمار(تشکیل پرونده شده)",
    "نام بيمار(تشكيل پرونده شده)", "نام", "نام خانوادگی",
]
PHONE_HEADERS = ["موبایل", "موبايل", "تلفن", "شماره تماس", "تلفن همراه", "شماره موبایل"]
DATE_HEADERS = ["تاریخ", "تاريخ", "تاریخ نوبت", "تاریخ ویزیت", "تاریخ مراجعه", "تاريخ نوبت"]
TIME_HEADERS = ["زمان", "ساعت", "وقت"]
RECORD_NO_HEADERS = ["پرونده", "شماره پرونده", "کد پرونده", "record_no", "کد بیمار"]
DOCTOR_HEADERS = ["نام پزشک", "پزشک", "دکتر", "نام دکتر"]
SERVICE_HEADERS = ["خدمات", "خدمت", "نام خدمت", "نوع خدمت"]
INSURANCE_HEADERS = ["بیمه", "بيمه", "سازمان بیمه"]
GENDER_HEADERS = ["جنسیت", "جنس"]
NOTES_HEADERS = ["توضیحات", "یادداشت", "ملاحظات"]


def _norm_header(h) -> str:
    if h is None:
        return ""
    t = str(h).strip().replace("\u200c", " ").replace("ي", "ی").replace("ك", "ک")
    return " ".join(t.split())


def _find_col(header_norm: dict, candidates: list) -> str | None:
    for c in candidates:
        n = _norm_header(c)
        if n in header_norm:
            return header_norm[n]
        for k in header_norm:
            if n in k or k in n:
                return header_norm[k]
    return None


def _safe_str(val) -> str | None:
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        return None
    s = str(val).strip()
    return s if s else None


def run_schema(conn) -> None:
    for name in ("001_identity_resolution_schema.sql", "002_identity_resolution_indexes.sql"):
        path = SQL_DIR / name
        if path.exists():
            conn.executescript(path.read_text(encoding="utf-8"))
    conn.commit()


def get_column_map(headers: list[str]) -> dict:
    norm = {_norm_header(h): h for h in headers}
    return {
        "name": _find_col(norm, NAME_HEADERS),
        "name_last": _find_col(norm, ["نام خانوادگی", "نام خانوادگی"]) or _find_col(norm, NAME_HEADERS),
        "phone": _find_col(norm, PHONE_HEADERS),
        "date": _find_col(norm, DATE_HEADERS),
        "time": _find_col(norm, TIME_HEADERS),
        "record_no": _find_col(norm, RECORD_NO_HEADERS),
        "doctor": _find_col(norm, DOCTOR_HEADERS),
        "service": _find_col(norm, SERVICE_HEADERS),
        "insurance": _find_col(norm, INSURANCE_HEADERS),
        "gender": _find_col(norm, GENDER_HEADERS),
        "notes": _find_col(norm, NOTES_HEADERS),
    }


def import_one_file(conn, path: Path, source_file: str, shamsi_year: int) -> dict:
    from openpyxl import load_workbook

    if not path.exists():
        return {"file": source_file, "error": "File not found", "inserted": 0}

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows)
        wb.close()
    except Exception as e:
        return {"file": source_file, "error": str(e), "inserted": 0}

    headers = [str(c).strip() if c is not None else "" for c in header_row]
    col = get_column_map(headers)
    if not col.get("name"):
        return {"file": source_file, "error": "No patient name column found", "inserted": 0}

    conn.execute("DELETE FROM appointments_unified_staging WHERE source_file = ?", (source_file,))
    conn.commit()

    ins = """
        INSERT INTO appointments_unified_staging (
            source_file, shamsi_year, sheet_name, source_row_number,
            appointment_date_raw, appointment_time_raw,
            patient_name_raw, patient_last_name_raw, patient_name_combined_raw,
            phone_raw, insurance_raw, doctor_name_raw, service_name_raw,
            gender_raw, notes_raw, appointment_type_raw, record_no_raw, created_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    created_at = datetime.now().isoformat()
    batch = []
    row_num = 1
    inserted = 0

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        sheet_name = ws.title or "Sheet1"
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            def cell(key):
                h = col.get(key)
                if h is None:
                    return None
                try:
                    idx = list(headers).index(h) if h in headers else None
                except ValueError:
                    idx = None
                if idx is not None and idx < len(row):
                    return row[idx]
                return None

            date_raw = _safe_str(cell("date"))
            time_raw = _safe_str(cell("time"))
            name_raw = _safe_str(cell("name"))
            name_last_raw = _safe_str(cell("name_last"))
            combined = name_raw or ""
            if name_last_raw and name_last_raw != name_raw:
                combined = f"{combined} {name_last_raw}".strip()
            if not combined:
                combined = name_raw

            batch.append((
                source_file, shamsi_year, sheet_name, row_num,
                date_raw, time_raw,
                name_raw, name_last_raw, combined,
                _safe_str(cell("phone")),
                _safe_str(cell("insurance")),
                _safe_str(cell("doctor")),
                _safe_str(cell("service")),
                _safe_str(cell("gender")),
                _safe_str(cell("notes")),
                None,
                _safe_str(cell("record_no")),
                created_at,
            ))
            if len(batch) >= BATCH_SIZE:
                conn.executemany(ins, batch)
                conn.commit()
                inserted += len(batch)
                batch.clear()
        if batch:
            conn.executemany(ins, batch)
            conn.commit()
            inserted += len(batch)
        wb.close()
    except Exception as e:
        conn.rollback()
        return {"file": source_file, "error": str(e), "inserted": inserted}
    return {"file": source_file, "inserted": inserted, "year": shamsi_year}


def main():
    import sqlite3
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    db_path = os.environ.get("ATIEH_DB_PATH") or os.environ.get("DB_PATH") or str(REPO / "atieh_clinic_recovery81_test.db")
    db_path = Path(db_path)
    if not db_path.is_absolute():
        db_path = REPO / db_path
    if not db_path.exists():
        print(f"Database not found: {db_path}", file=sys.stderr)
        return 1

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA busy_timeout = 30000")
    run_schema(conn)

    print(f"Unified appointments import -> {db_path}")
    print(f"Staging table: appointments_unified_staging")
    print(f"History base: {HISTORY_BASE}")
    total = 0
    for rel, year in APPOINTMENT_FILE_PATTERNS:
        path = REPO / "data" / "inputs" / "history" / rel
        if not path.exists():
            path = HISTORY_BASE / rel
        if not path.exists():
            path = REPO / rel
        name = Path(rel).name
        r = import_one_file(conn, path, name, year)
        safe_name = name.encode("ascii", "replace").decode("ascii") if name else str(r.get("year", ""))
        if r.get("error"):
            print(f"  {safe_name}: ERROR {r['error']}")
        else:
            print(f"  {safe_name}: {r['inserted']} rows")
            total += r["inserted"]
    print(f"\nTotal inserted: {total}")
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
