# -*- coding: utf-8 -*-
"""
Unified Payments Import – یک importer مشترک برای همه فایل‌های payments سالانه.
- Column mapping ثابت (موبايل، كد ملي، خالص دريافتي، نام بيمار، تاريخ پذيرش)
- record_no اختیاری: اگر ستون شماره پرونده وجود داشت بخوان، وگرنه NULL
- برای هر رکورد year و source_file ذخیره می‌شود
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from datetime import datetime

REPO = Path(__file__).resolve().parent.parent
PAYMENTS_DIR = REPO / "data" / "inputs" / "payments"
SQL_DIR = REPO / "sql"
BATCH_SIZE = 2000

# Fixed column name candidates (same as audit)
COL_CANDIDATES = {
    "mobile": ["موبایل", "موبايل", "تلفن", "شماره تماس"],
    "national_id": ["کد ملی", "كد ملي", "کد ملي", "کدملی"],
    "net_received": ["خالص دریافتی", "خالص دريافتي"],
    "patient_name": ["نام بیمار", "نام بيمار", "نام بیمار(تشکیل پرونده شده)"],
    "record_no": ["شماره پرونده", "کد پرونده", "record_no"],
    "admission_date": ["تاریخ پذیرش", "تاريخ پذيرش"],
    "insurer": ["سازمان بیمه گر بیمار", "سازمان |بيمه گر بيمار", "بیمه گر"],
    "amount_patient": ["سهم بیمار", "سهم بيمار"],
    "amount_insurer": ["سهم سازمان", "سهم سازمان"],
}


def _norm(s) -> str:
    if s is None:
        return ""
    t = str(s).strip().replace("ي", "ی").replace("ك", "ک").replace("|", " ")
    return " ".join(t.split())


def find_column(col_norm: dict, candidates: list) -> int | None:
    for c in candidates:
        if c in col_norm:
            return col_norm[c]
    for c in candidates:
        for k in col_norm:
            if c in k or k in c:
                return col_norm[k]
    return None


def _safe_str(val) -> str | None:
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        return None
    s = str(val).strip()
    return s if s else None


def extract_record_no_from_patient_name(patient_name_raw: str | None) -> str | None:
    """Optional helper: استخراج شماره پرونده از نام بیمار مثلاً 'نام(12345)' → '12345'. در importer استفاده نمی‌شود."""
    if not patient_name_raw or not patient_name_raw.strip():
        return None
    m = re.search(r"\((\d{4,10})\)\s*$", patient_name_raw.strip())
    return m.group(1) if m else None


def _extract_year_from_filename(name: str) -> int | None:
    m = re.search(r"payments_(\d{4})_full", name, re.IGNORECASE)
    return int(m.group(1)) if m else None


def run_schema(conn) -> None:
    schema_path = SQL_DIR / "unified_payments_staging_schema.sql"
    if not schema_path.exists():
        raise FileNotFoundError(f"Schema not found: {schema_path}")
    conn.executescript(schema_path.read_text(encoding="utf-8"))


def get_column_indices(headers: list[str]) -> dict[str, int | None]:
    col_norm = {_norm(h): i for i, h in enumerate(headers)}
    return {key: find_column(col_norm, cands) for key, cands in COL_CANDIDATES.items()}


def import_one_file(conn, path: Path, *, truncate_first: bool = False) -> dict:
    """Import one payments_*_full.xlsx into payments_unified_staging."""
    from openpyxl import load_workbook

    name = path.name
    year = _extract_year_from_filename(name)
    if year is None:
        return {"file": name, "error": "Could not extract year from filename", "inserted": 0}

    if truncate_first:
        conn.execute("DELETE FROM payments_unified_staging WHERE source_file = ?", (name,))
        conn.commit()

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows)
        wb.close()
    except Exception as e:
        return {"file": name, "error": str(e), "inserted": 0}

    headers = [str(c).strip() if c is not None else "" for c in header_row]
    idx = get_column_indices(headers)
    if idx.get("patient_name") is None:
        return {"file": name, "error": "Required column patient_name not found", "inserted": 0}

    loaded_at = datetime.now().isoformat()
    insert_sql = """
        INSERT INTO payments_unified_staging (
            source_file, shamsi_year, row_number, sheet_name, loaded_at,
            parse_status, parse_error,
            patient_name_raw, phone_raw, national_id_raw, net_received_raw,
            record_no, appointment_date_raw, insurer_raw, amount_patient_raw, amount_insurer_raw
        ) VALUES (?,?,?,?,?, 'ok', NULL, ?,?,?,?, ?,?,?,?,?)
    """
    batch = []
    row_num = 1
    inserted = 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            def cell(k):
                i = idx.get(k)
                return row[i] if i is not None and i < len(row) else None
            patient_name_raw = _safe_str(cell("patient_name"))
            phone_raw = _safe_str(cell("phone"))
            national_id_raw = _safe_str(cell("national_id"))
            net_received_raw = _safe_str(cell("net_received"))
            record_no = _safe_str(cell("record_no"))  # optional
            appointment_date_raw = _safe_str(cell("admission_date"))
            insurer_raw = _safe_str(cell("insurer"))
            amount_patient_raw = _safe_str(cell("amount_patient"))
            amount_insurer_raw = _safe_str(cell("amount_insurer"))

            batch.append((
                name, year, row_num, ws.title or None, loaded_at,
                patient_name_raw, phone_raw, national_id_raw, net_received_raw,
                record_no, appointment_date_raw, insurer_raw, amount_patient_raw, amount_insurer_raw,
            ))
            if len(batch) >= BATCH_SIZE:
                conn.executemany(insert_sql, batch)
                conn.commit()
                inserted += len(batch)
                batch.clear()
        if batch:
            conn.executemany(insert_sql, batch)
            conn.commit()
            inserted += len(batch)
        wb.close()
    except Exception as e:
        conn.rollback()
        return {"file": name, "error": str(e), "inserted": inserted}
    return {"file": name, "inserted": inserted, "year": year}


def main():
    import sqlite3
    import os

    db_path = os.environ.get("ATIEH_DB_PATH") or os.environ.get("DB_PATH")
    if not db_path:
        db_path = REPO / "atieh_clinic.db"
    db_path = Path(db_path)
    if not db_path.is_absolute():
        db_path = REPO / db_path

    if not db_path.exists():
        print(f"Database not found: {db_path}", file=sys.stderr)
        return 1

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA busy_timeout = 30000")
    run_schema(conn)

    files = sorted(PAYMENTS_DIR.glob("payments_*_full.xlsx"))
    files = [f for f in files if not f.name.startswith("~$")]
    if not files:
        print("No payment files found.", file=sys.stderr)
        return 1

    print(f"Unified import into: {db_path}")
    print(f"Staging table: payments_unified_staging")
    print(f"Files: {len(files)}")
    for f in files:
        print(f"  - {f.name}")
    print()

    total = 0
    for path in files:
        r = import_one_file(conn, path, truncate_first=True)
        if r.get("error"):
            print(f"  {r['file']}: ERROR {r['error']}")
        else:
            print(f"  {r['file']}: {r['inserted']} rows")
            total += r["inserted"]
    print(f"\nTotal inserted: {total}")
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
