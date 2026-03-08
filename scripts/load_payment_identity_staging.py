from pathlib import Path
import sqlite3
import re
from openpyxl import load_workbook

DB = r".\atieh_clinic_working.db"
PAYMENTS_DIR = Path(r".\data\inputs\payments")

FILES = [
    PAYMENTS_DIR / "payments_1395_full.xlsx",
    PAYMENTS_DIR / "payments_1396_full.xlsx",
    PAYMENTS_DIR / "payments_1397_full.xlsx",
    PAYMENTS_DIR / "payments_1398_full.xlsx",
    PAYMENTS_DIR / "payments_1399_full.xlsx",
    PAYMENTS_DIR / "payments_1400_full.xlsx",
    PAYMENTS_DIR / "payments_1402_full.xlsx",
    PAYMENTS_DIR / "payments_1403_full.xlsx",
    PAYMENTS_DIR / "payments_1404_full.xlsx",
]

TARGET_TABLE = "payment_identity_staging"

def normalize_text(x):
    if x is None:
        return ""
    s = str(x).strip()
    s = s.replace("\u200c", " ").replace("\u200f", " ").replace("\ufeff", " ")
    s = s.replace("ي", "ی").replace("ك", "ک")
    s = s.strip()
    while len(s) >= 2 and s[0] in "\"'`" and s[-1] in "\"'`":
        s = s[1:-1].strip()
    s = s.strip("\"'` ")
    s = re.sub(r"\s+", " ", s)
    return s

def to_str(x):
    if x is None:
        return None
    s = str(x).strip()
    return s if s != "" else None

def get_header_map(header_row):
    norm_headers = [normalize_text(h) for h in header_row]
    mapping = {}
    for i, h in enumerate(norm_headers):
        mapping[h] = i
    return norm_headers, mapping

def find_required_columns(norm_headers):
    aliases = {
        "receipt_no": [
            "شماره رسید",
            "شماره رسيد",
        ],
        "record_no": [
            "شماره پرونده",
        ],
        "patient_name_raw": [
            "نام بیمار",
            "نام بيمار",
        ],
        "mobile_raw": [
            "موبایل",
            "موبايل",
        ],
        "national_id_raw": [
            "کد ملی",
            "كد ملي",
            "کد ملي",
        ],
        "admission_date_raw": [
            "تاریخ پذیرش",
            "تاريخ پذيرش",
        ],
        "net_received_raw": [
            "خالص دریافتی",
            "خالص دريافتي",
        ],
    }

    found = {}
    for key, options in aliases.items():
        idx = None
        for i, h in enumerate(norm_headers):
            if h in options:
                idx = i
                break
        found[key] = idx
    return found

def main():
    print("Files selected for staging:")
    for f in FILES:
        print(" -", f)

    conn = sqlite3.connect(DB, timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000;")
    conn.execute("PRAGMA journal_mode = WAL;")
    conn.execute("PRAGMA synchronous = NORMAL;")
    cur = conn.cursor()

    print(f"Clearing table: {TARGET_TABLE}")
    cur.execute(f"DELETE FROM {TARGET_TABLE}")
    conn.commit()

    insert_sql = f'''
    INSERT INTO {TARGET_TABLE}
    (
        source_file,
        receipt_no,
        record_no,
        patient_name_raw,
        mobile_raw,
        national_id_raw,
        admission_date_raw,
        net_received_raw
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    '''

    grand_total = 0

    for file_path in FILES:
        if not file_path.exists():
            print(f"[SKIP] File not found: {file_path}")
            continue

        print(f"\n=== Processing: {file_path.name} ===")
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name}")

        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            print("[SKIP] Empty sheet")
            continue

        norm_headers, header_map = get_header_map(header_row)
        print("Detected normalized headers:")
        print(norm_headers)

        cols = find_required_columns(norm_headers)
        print("Column positions:", cols)

        missing = [k for k, v in cols.items() if v is None]
        if missing:
            print(f"[ERROR] Missing required columns in {file_path.name}: {missing}")
            continue

        batch = []
        file_rows_seen = 0
        file_rows_inserted = 0

        for row in rows:
            file_rows_seen += 1

            def val(col_key):
                idx = cols[col_key]
                if idx is None or idx >= len(row):
                    return None
                return to_str(row[idx])

            record_no = val("record_no")
            patient_name_raw = val("patient_name_raw")
            mobile_raw = val("mobile_raw")
            national_id_raw = val("national_id_raw")
            receipt_no = val("receipt_no")
            admission_date_raw = val("admission_date_raw")
            net_received_raw = val("net_received_raw")

            if not any([record_no, patient_name_raw, mobile_raw, national_id_raw, receipt_no, admission_date_raw, net_received_raw]):
                continue

            batch.append((
                file_path.name,
                receipt_no,
                record_no,
                patient_name_raw,
                mobile_raw,
                national_id_raw,
                admission_date_raw,
                net_received_raw
            ))

            if len(batch) >= 1000:
                cur.executemany(insert_sql, batch)
                conn.commit()
                file_rows_inserted += len(batch)
                grand_total += len(batch)
                print(f"  inserted={file_rows_inserted} seen={file_rows_seen}")
                batch = []

        if batch:
            cur.executemany(insert_sql, batch)
            conn.commit()
            file_rows_inserted += len(batch)
            grand_total += len(batch)

        print(f"[DONE] {file_path.name}: seen={file_rows_seen}, inserted={file_rows_inserted}")

    print(f"\nGrand total inserted: {grand_total}")
    conn.close()

if __name__ == "__main__":
    main()
