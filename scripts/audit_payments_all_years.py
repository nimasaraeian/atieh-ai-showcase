# -*- coding: utf-8 -*-
"""
Audit ساختاری همه فایل‌های payments (همه سال‌ها).
خروجی: گزارش per-year + گزارش تجمیعی مقایسه‌ای.
بدون import/recovery — فقط audit و report.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from dataclasses import dataclass, field

REPO = Path(__file__).resolve().parent.parent
PAYMENTS_DIR = REPO / "data" / "inputs" / "payments"
SAMPLE_ROWS = 3_000  # سطر نمونه برای آمار (برای سرعت)


def _norm(s) -> str:
    if s is None:
        return ""
    t = str(s).strip().replace("ي", "ی").replace("ك", "ک").replace("|", " ")
    return " ".join(t.split())


# نام‌های ممکن برای هر ستون کلیدی (همه گونه‌های ی/ک و فاصله)
COL_CANDIDATES = {
    "mobile": ["موبایل", "موبايل", "تلفن", "شماره تماس", "موبايل"],
    "national_id": ["کد ملی", "كد ملي", "کد ملي", "کدملی"],
    "net_received": ["خالص دریافتی", "خالص دريافتي", "خالص دريافتی"],
    "patient_name": ["نام بیمار", "نام بيمار", "نام بیمار(تشکیل پرونده شده)"],
    "record_no": ["شماره پرونده", "کد پرونده", "شماره پرونده", "record_no"],
    "admission_date": ["تاریخ پذیرش", "تاريخ پذيرش", "تاریخ پذیرش"],
}


def find_column(col_norm: dict, candidates: list) -> int | None:
    for c in candidates:
        if c in col_norm:
            return col_norm[c]
    for c in candidates:
        for k in col_norm:
            if c in k or k in c:
                return col_norm[k]
    return None


def is_valid_mobile(s: str) -> bool:
    digits = re.sub(r"\D", "", s)
    if digits.startswith("98"):
        digits = "0" + digits[2:]
    if len(digits) == 10 and digits.startswith("9"):
        digits = "0" + digits
    return len(digits) == 11 and digits.startswith("09")


def is_valid_national_id_10(s: str) -> bool:
    d = re.sub(r"\D", "", s)
    return len(d) == 10


def to_num(val) -> float | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("،", "")
    if not s or s in ("-", "—", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


@dataclass
class FileAudit:
    file_name: str
    year: str
    sheet_names: list[str] = field(default_factory=list)
    row_count_sample: int = 0
    col_count: int = 0
    col_found: dict = field(default_factory=dict)  # key -> actual header name or None
    col_index: dict = field(default_factory=dict)  # key -> 0-based index
    mobile_filled: int = 0
    mobile_valid: int = 0
    national_id_filled: int = 0
    national_id_valid10: int = 0
    net_received_numeric: int = 0
    net_received_nonzero: int = 0
    record_no_exists: bool = False
    record_no_filled: int = 0
    samples: dict = field(default_factory=dict)  # key -> list of up to 5 sample values
    error: str | None = None


def audit_one_file(path: Path) -> FileAudit:
    name = path.name
    year = name.replace("payments_", "").replace("_full.xlsx", "").replace(".xlsx", "")
    audit = FileAudit(file_name=name, year=year)

    try:
        from openpyxl import load_workbook
    except ImportError:
        audit.error = "openpyxl not installed"
        return audit

    if not path.exists():
        audit.error = "File not found"
        return audit

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        audit.sheet_names = wb.sheetnames
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=SAMPLE_ROWS + 1, values_only=True))
        wb.close()
    except Exception as e:
        audit.error = str(e)
        return audit

    if not rows:
        audit.error = "No rows"
        return audit

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    audit.col_count = len(headers)
    audit.row_count_sample = len(rows) - 1

    col_norm = {_norm(h): i for i, h in enumerate(headers)}

    for key, candidates in COL_CANDIDATES.items():
        idx = find_column(col_norm, candidates)
        audit.col_index[key] = idx
        audit.col_found[key] = headers[idx] if idx is not None else None
        if key == "record_no":
            audit.record_no_exists = idx is not None

    data_rows = rows[1:]

    # Mobile
    idx = audit.col_index.get("mobile")
    for r in data_rows:
        v = r[idx] if idx is not None else None
        s = (str(v).strip() if v is not None and str(v).strip() else "")
        if s:
            audit.mobile_filled += 1
            if is_valid_mobile(s):
                audit.mobile_valid += 1
            if len(audit.samples.setdefault("mobile", [])) < 5:
                audit.samples["mobile"].append(s)

    # National ID
    idx = audit.col_index.get("national_id")
    for r in data_rows:
        v = r[idx] if idx is not None else None
        s = (str(v).strip() if v is not None and str(v).strip() else "")
        if s:
            audit.national_id_filled += 1
            if is_valid_national_id_10(s):
                audit.national_id_valid10 += 1
            if len(audit.samples.setdefault("national_id", [])) < 5:
                audit.samples["national_id"].append(s)

    # Net received
    idx = audit.col_index.get("net_received")
    for r in data_rows:
        v = r[idx] if idx is not None else None
        x = to_num(v)
        if x is not None:
            audit.net_received_numeric += 1
            if x != 0:
                audit.net_received_nonzero += 1
            if len(audit.samples.setdefault("net_received", [])) < 5:
                audit.samples["net_received"].append(x)

    # Record_no filled
    idx = audit.col_index.get("record_no")
    if idx is not None:
        audit.record_no_filled = sum(
            1 for r in data_rows
            if r[idx] is not None and str(r[idx]).strip()
        )
        for r in data_rows:
            if r[idx] is not None and str(r[idx]).strip():
                if len(audit.samples.setdefault("record_no", [])) < 5:
                    audit.samples["record_no"].append(str(r[idx]).strip())
                break
    else:
        audit.samples["record_no"] = []

    # Patient name samples
    idx = audit.col_index.get("patient_name")
    if idx is not None:
        for r in data_rows:
            v = r[idx]
            if v is not None and str(v).strip():
                if len(audit.samples.setdefault("patient_name", [])) < 5:
                    audit.samples["patient_name"].append(str(v).strip()[:50])
                break

    # Admission date samples
    idx = audit.col_index.get("admission_date")
    if idx is not None:
        for r in data_rows:
            v = r[idx]
            if v is not None and str(v).strip():
                if len(audit.samples.setdefault("admission_date", [])) < 5:
                    audit.samples["admission_date"].append(str(v).strip())
                break

    return audit


def run_all_audits() -> list[FileAudit]:
    files = sorted(PAYMENTS_DIR.glob("payments_*_full.xlsx"))
    files = [f for f in files if not f.name.startswith("~$")]
    results = []
    for path in files:
        results.append(audit_one_file(path))
    return results


def build_report(audits: list[FileAudit]) -> str:
    lines = []

    lines.append("# گزارش تجمیعی Audit ساختاری فایل‌های Payments (همه سال‌ها)")
    lines.append("")
    lines.append("**هدف:** تحلیل ساختاری بدون import/recovery. تعیین ستون‌های کلیدی، پوشش موبایل/کد ملی/خالص دریافتی، و امکان importer مشترک یا per-year.")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ─── بخش ۱: هر سال ───
    lines.append("## ۱. جزئیات هر سال")
    lines.append("")

    for a in audits:
        lines.append(f"### {a.file_name} (سال {a.year})")
        lines.append("")
        if a.error:
            lines.append(f"- **خطا:** {a.error}")
            lines.append("")
            continue

        lines.append("#### ۱.۱ شیت‌ها و ابعاد")
        lines.append(f"- نام شیت‌ها: {', '.join(a.sheet_names)}")
        lines.append(f"- تعداد سطر (نمونه): **{a.row_count_sample}**")
        lines.append(f"- تعداد ستون: **{a.col_count}**")
        lines.append("")

        lines.append("#### ۱.۲ نام واقعی ستون‌های کلیدی")
        lines.append("| کلید | نام ستون در فایل |")
        lines.append("|------|-------------------|")
        for key in ["mobile", "national_id", "net_received", "patient_name", "record_no", "admission_date"]:
            name = a.col_found.get(key) or "—"
            lines.append(f"| {key} | {name} |")
        lines.append("")

        lines.append("#### ۱.۳ آمار نمونه")
        lines.append(f"- موبایل پر: **{a.mobile_filled}**")
        lines.append(f"- موبایل معتبر (بعد از نرمال): **{a.mobile_valid}**")
        lines.append(f"- کد ملی پر: **{a.national_id_filled}**")
        lines.append(f"- کد ملی معتبر ۱۰ رقمی: **{a.national_id_valid10}**")
        lines.append(f"- خالص دریافتی عددی: **{a.net_received_numeric}**")
        lines.append(f"- خالص دریافتی غیرصفر: **{a.net_received_nonzero}**")
        lines.append(f"- وجود ستون record_no: **{'بله' if a.record_no_exists else 'خیر'}**")
        if a.record_no_exists:
            lines.append(f"- record_no پر: **{a.record_no_filled}**")
        lines.append("")

        lines.append("#### ۱.۴ نمونه مقادیر")
        for key in ["mobile", "national_id", "net_received", "patient_name", "record_no", "admission_date"]:
            samples = a.samples.get(key, [])
            if samples:
                lines.append(f"- **{key}:** " + ", ".join(repr(s) for s in samples[:5]))
            else:
                lines.append(f"- **{key}:** —")
        lines.append("")
        lines.append("---")
        lines.append("")

    # ─── بخش ۲: جدول مقایسه ───
    lines.append("## ۲. جدول مقایسه‌ای بین سال‌ها")
    lines.append("")

    lines.append("| سال | سطر(نمونه) | ستون | موبایل پر | موبایل معتبر | کدملی پر | کدملی۱۰ | net عددی | net≠0 | record_no ستون | record_no پر |")
    lines.append("|-----|-------------|------|-----------|--------------|----------|---------|----------|-------|----------------|--------------|")
    for a in audits:
        if a.error:
            lines.append(f"| {a.year} | خطا | — | — | — | — | — | — | — | — | — |")
            continue
        rec_col = "بله" if a.record_no_exists else "خیر"
        rec_fill = str(a.record_no_filled) if a.record_no_exists else "—"
        lines.append(
            f"| {a.year} | {a.row_count_sample} | {a.col_count} | {a.mobile_filled} | {a.mobile_valid} | "
            f"{a.national_id_filled} | {a.national_id_valid10} | {a.net_received_numeric} | {a.net_received_nonzero} | {rec_col} | {rec_fill} |"
        )
    lines.append("")

    # ─── بخش ۳: تفاوت ساختاری ───
    lines.append("## ۳. تفاوت ساختاری بین سال‌ها")
    lines.append("")

    # Column name differences
    by_col = {}
    for key in ["mobile", "national_id", "net_received", "patient_name", "record_no", "admission_date"]:
        by_col[key] = {}
        for a in audits:
            if a.error:
                continue
            name = a.col_found.get(key) or "(ندارد)"
            by_col[key][a.year] = name

    lines.append("### ۳.۱ نام ستون‌های کلیدی به تفکیک سال")
    lines.append("")
    for key in ["mobile", "national_id", "net_received", "patient_name", "record_no", "admission_date"]:
        lines.append(f"- **{key}:**")
        for a in audits:
            if a.error:
                continue
            name = by_col[key].get(a.year, "—")
            lines.append(f"  - {a.year}: `{name}`")
        lines.append("")

    # Structural: same or different (keep "(ندارد)" so missing column is visible)
    unique_headers = {}
    for key in ["mobile", "national_id", "net_received", "patient_name", "record_no", "admission_date"]:
        names = set(v or "(ندارد)" for v in by_col[key].values())
        unique_headers[key] = sorted(names)

    lines.append("### ۳.۲ یکسان یا متفاوت بودن نام ستون‌ها")
    lines.append("")
    for key, names in unique_headers.items():
        if len(names) <= 1 and "(ندارد)" not in names:
            lines.append(f"- **{key}:** در همه سال‌ها یکسان است.")
        elif "(ندارد)" in names and len(names) == 2:
            other = [n for n in names if n != "(ندارد)"][0]
            lines.append(f"- **{key}:** متفاوت — در یک یا چند سال ستون جدا وجود ندارد (بقیه: `{other}`).")
        elif len(names) > 1:
            lines.append(f"- **{key}:** متفاوت — گونه‌های یافت‌شده: {names}")
        else:
            lines.append(f"- **{key}:** گونه‌ها: {names}")
    lines.append("")

    # ─── بخش ۴: توصیه‌ها ───
    lines.append("## ۴. جمع‌بندی و توصیه‌ها")
    lines.append("")

    lines.append("### ۴.۱ کدام سال‌ها برای National ID Recovery مناسب‌ترند؟")
    lines.append("")
    nid_ok = [(a.year, a.national_id_valid10, a.row_count_sample) for a in audits if not a.error and a.row_count_sample]
    nid_ok.sort(key=lambda x: (x[1] / x[2] if x[2] else 0, x[1]), reverse=True)
    for year, valid, total in nid_ok:
        pct = (100 * valid / total) if total else 0
        lines.append(f"- **{year}:** {valid} کد ملی ۱۰ رقمی معتبر از {total} سطر ({pct:.1f}%)")
    lines.append("")
    if nid_ok:
        best_nid = nid_ok[0][0]
        lines.append(f"**توصیه:** سال‌های با بیشترین نسبت کد ملی معتبر برای NID recovery در اولویت: بر اساس نسبت و تعداد مطلق، **{best_nid}** و سال‌های با درصد مشابه مناسب‌ترند.")
    lines.append("")

    lines.append("### ۴.۲ کدام سال‌ها برای Phone Recovery مناسب‌ترند؟")
    lines.append("")
    phone_ok = [(a.year, a.mobile_valid, a.row_count_sample) for a in audits if not a.error and a.row_count_sample]
    phone_ok.sort(key=lambda x: (x[1] / x[2] if x[2] else 0, x[1]), reverse=True)
    for year, valid, total in phone_ok:
        pct = (100 * valid / total) if total else 0
        lines.append(f"- **{year}:** {valid} موبایل معتبر از {total} سطر ({pct:.1f}%)")
    lines.append("")
    if phone_ok:
        best_phone = phone_ok[0][0]
        lines.append(f"**توصیه:** سال‌های با بیشترین پوشش موبایل معتبر برای Phone recovery: **{best_phone}** و سال‌های با درصد مشابه.")
    lines.append("")

    lines.append("### ۴.۳ آیا می‌توان یک Importer/Recovery مشترک برای همه سال‌ها ساخت؟")
    lines.append("")
    same_structure = all(
        len(set(by_col[k].get(a.year) or "(ندارد)" for a in audits if not a.error)) <= 1
        for k in ["mobile", "national_id", "net_received", "patient_name", "admission_date"]
    )
    record_no_mixed = any(a.record_no_exists for a in audits if not a.error) and any(not a.record_no_exists for a in audits if not a.error)

    if same_structure and not record_no_mixed:
        lines.append("**بله.** نام ستون‌های کلیدی (موبایل، کد ملی، خالص دریافتی، نام بیمار، تاریخ پذیرش) در سال‌ها یکسان یا سازگار است. می‌توان یک **importer مشترک** با یک نقشه ستون (با چند نام جایگزین per column) برای همه سال‌ها استفاده کرد.")
    else:
        lines.append("**نیاز به نقشه per-year یا منطق شرطی:**")
        if not same_structure:
            lines.append("- نام برخی ستون‌ها بین سال‌ها متفاوت است؛ باید در کد یک **جدول نقشه (year → column names)** یا تشخیص خودکار هدر داشته باشید.")
        if record_no_mixed:
            lines.append("- ستون **record_no** در بعضی سال‌ها وجود دارد و در بعضی خیر؛ در importer/recovery باید وجود این ستون را به‌صورت اختیاری چک کنید (یا record_no را از داخل نام بیمار استخراج کنید).")
    lines.append("")
    lines.append("**جمع‌بندی نهایی:** بر اساس این audit، اگر نام ستون‌ها با نرمال‌سازی (ی/ک و فاصله) و چند نام جایگزین (موبايل/موبایل، كد ملي/کد ملی، ...) پوشش داده شوند، یک **importer مشترک با column mapping قابل تنظیم** برای همه سال‌ها امکان‌پذیر است؛ در غیر این صورت **per-year mapping** امن‌تر است.")
    lines.append("")

    return "\n".join(lines)


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    audits = run_all_audits()
    report = build_report(audits)
    out_path = REPO / "docs" / "reports" / "payments_all_years_structural_audit.md"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(report, encoding="utf-8")
    print(f"Report written: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
