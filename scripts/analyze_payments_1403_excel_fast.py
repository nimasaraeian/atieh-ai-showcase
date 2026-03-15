# -*- coding: utf-8 -*-
"""تحلیل سریع اکسل ۱۴۰۳ با openpyxl read_only - فقط ۲۵۰۰ سطر اول."""
from __future__ import annotations

import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO / "data" / "inputs" / "payments" / "payments_1403_full.xlsx"


def _norm(s):
    if s is None:
        return ""
    t = str(s).strip().replace("ي", "ی").replace("ك", "ک")
    return " ".join(t.split())


def main():
    from openpyxl import load_workbook

    if not XLSX_PATH.exists():
        print("File not found", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=2501, values_only=True))
    wb.close()

    if not rows:
        print("No rows")
        return 0

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_norm = {_norm(h): i for i, h in enumerate(headers)}

    def find(*candidates):
        for c in candidates:
            if c in col_norm:
                return col_norm[c]
        for c in candidates:
            for k in col_norm:
                if c in k or k in c:
                    return col_norm[k]
        return None

    idx_phone = find("موبایل", "موبايل", "تلفن")
    idx_nid = find("کد ملی", "كد ملي", "کد ملي")
    idx_net = find("خالص دریافتی", "خالص دريافتي")
    idx_name = find("نام بیمار", "نام بيمار")
    idx_record = find("شماره پرونده", "کد پرونده")
    idx_date = find("تاریخ پذیرش", "تاريخ پذيرش")

    data_rows = rows[1:]
    n = len(data_rows)

    # Phone
    phone_filled = 0
    phone_valid = 0
    phone_samples = []
    for r in data_rows:
        v = r[idx_phone] if idx_phone is not None else None
        s = (str(v).strip() if v is not None and str(v).strip() else "")
        if s:
            phone_filled += 1
            digits = re.sub(r"\D", "", s)
            if digits.startswith("98"):
                digits = "0" + digits[2:]
            if len(digits) == 10 and digits.startswith("9"):
                digits = "0" + digits
            if len(digits) == 11 and digits.startswith("09"):
                phone_valid += 1
            if len(phone_samples) < 8:
                phone_samples.append(s)

    # National ID
    nid_filled = 0
    nid_valid10 = 0
    nid_samples = []
    for r in data_rows:
        v = r[idx_nid] if idx_nid is not None else None
        s = (str(v).strip() if v is not None and str(v).strip() else "")
        if s:
            nid_filled += 1
            d = re.sub(r"\D", "", s)
            if len(d) == 10:
                nid_valid10 += 1
            if len(nid_samples) < 8:
                nid_samples.append(s)

    # Net received
    net_filled = 0
    net_sum = 0
    net_vals = []
    net_min = None
    net_max = None
    for r in data_rows:
        v = r[idx_net] if idx_net is not None else None
        if v is None:
            continue
        s = str(v).strip().replace(",", "").replace("،", "")
        if not s or s in ("-", "nan"):
            continue
        try:
            x = float(s)
            net_filled += 1
            net_sum += x
            if net_min is None or x < net_min:
                net_min = x
            if net_max is None or x > net_max:
                net_max = x
            if len(net_vals) < 8:
                net_vals.append(x)
        except ValueError:
            pass

    # Name
    name_filled = sum(1 for r in data_rows if (r[idx_name] if idx_name is not None else None) and str(r[idx_name]).strip())

    # Record no
    rec_filled = sum(1 for r in data_rows if (r[idx_record] if idx_record is not None else None) and str(r[idx_record]).strip())

    report = []
    report.append("# گزارش کامل فایل اکسل payments_1403_full.xlsx")
    report.append("")
    report.append("## ۱. اطلاعات فایل")
    report.append("")
    report.append(f"- مسیر: `{XLSX_PATH.name}`")
    report.append(f"- تحلیل بر اساس **۲۵۰۰ سطر اول** شیت اول (برای سرعت).")
    report.append("")
    report.append("## ۲. ستون‌های شناسایی‌شده")
    report.append("")
    report.append("| ستون | نام یافت‌شده |")
    report.append("|------|-------------|")
    for label, idx in [("تلفن/موبایل", idx_phone), ("کد ملی", idx_nid), ("خالص دریافتی", idx_net), ("نام بیمار", idx_name), ("شماره پرونده", idx_record), ("تاریخ پذیرش", idx_date)]:
        name = headers[idx] if idx is not None else "—"
        report.append(f"| {label} | {name} |")
    report.append("")
    report.append("## ۳. لیست تمام ستون‌ها (هدر)")
    report.append("")
    for i, h in enumerate(headers):
        report.append(f"{i+1}. {h}")
    report.append("")
    report.append("## ۴. تحلیل ستون تلفن (موبایل)")
    report.append("")
    report.append(f"- تعداد سطرهای نمونه: **{n}**")
    report.append(f"- پر (غیر خالی): **{phone_filled}**")
    report.append(f"- موبایل معتبر (۱۰/۱۱ رقم، ۹ یا ۰۹): **{phone_valid}**")
    report.append("- در این فایل موبایل اغلب به صورت **۹xxxxxxxxx** (بدون صفر اول) ذخیره شده؛ برای تطابق با پایگاه باید با ۰۹ نرمال شود.")
    report.append("- نمونه:")
    for s in phone_samples:
        report.append(f"  - `{s}`")
    report.append("")
    report.append("## ۵. تحلیل ستون کد ملی")
    report.append("")
    report.append(f"- پر: **{nid_filled}**")
    report.append(f"- کد ۱۰ رقمی: **{nid_valid10}**")
    report.append("- نمونه:")
    for s in nid_samples:
        report.append(f"  - `{s}`")
    report.append("")
    report.append("## ۶. تحلیل ستون خالص دریافتی")
    report.append("")
    report.append(f"- پر (عددی): **{net_filled}**")
    report.append(f"- **جمع خالص دریافتی (نمونه):** {net_sum:,.0f}")
    if net_min is not None:
        report.append(f"- کمینه: **{net_min:,.0f}**")
    if net_max is not None:
        report.append(f"- بیشینه: **{net_max:,.0f}**")
    report.append("- نمونه مقادیر:")
    for v in net_vals:
        report.append(f"  - {v}")
    report.append("")
    report.append("## ۷. نام بیمار و شماره پرونده")
    report.append("")
    report.append(f"- نام بیمار پر: **{name_filled}**")
    report.append(f"- شماره پرونده (ستون جدا): پر **{rec_filled}** سطر.")
    if rec_filled == 0:
        report.append("- ستون جداگانه‌ای با نام «شماره پرونده» در هدر یافت نشد؛ احتمالاً شماره پرونده داخل نام بیمار به صورت `نام(عدد)` یا در ستون دیگری است.")
    report.append("")
    report.append("## ۸. خلاصه")
    report.append("")
    report.append("در این اکسل هر سه ستون **تلفن**، **کد ملی** و **خالص دریافتی** وجود دارند و قابل استفاده برای تطابق هویت و محاسبات مالی هستند.")
    report.append("")

    text = "\n".join(report)
    out_path = REPO / "docs" / "reports" / "payments_1403_full_excel_report.md"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")
    print("Report written to", str(out_path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
