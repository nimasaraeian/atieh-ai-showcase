import sqlite3
import openpyxl
import re
from pathlib import Path

DB_PATH = r"C:\Users\USER\Documents\GitHub\atieh\atieh_clinic_recovery81_test.db"
XLSX_PATH = r"C:\Users\USER\Documents\GitHub\atieh\data\inputs\reference\تاریخ پرداختی بیمه ها.xlsx"

def canon_insurer(name: str):
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None

    s = s.replace("ي","ی").replace("ك","ک")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\d+","", s)
    s = s.replace("%","").replace("("," ").replace(")"," ")
    s = s.replace("درصد"," ").replace("/", " ")
    s = re.sub(r"\s+", " ", s).strip()

    if "آزاد" in s or "نقد" in s:
        return "آزاد"
    if "تامين اجتماعي" in s or "تامین اجتماعی" in s:
        return "تامین اجتماعی"
    if "جانبازان نیروهای مسلح" in s:
        return "جانبازان نیروهای مسلح"
    if "نیروهای مسلح" in s:
        return "نیروهای مسلح"
    if "بانک ملت" in s:
        return "بانک ملت"
    if "بانک سپه" in s or s == "سپه":
        return "بانک سپه"
    if "بانک ملی" in s or s == "ملی":
        return "بانک ملی"
    if "کشاورزی" in s:
        return "بانک کشاورزی"
    if "شرکت نفت" in s:
        return "شرکت نفت"
    if "آسیا" in s or "اسیا" in s:
        return "آسیا"
    if s == "دی" or "بیمه دی" in s:
        return "دی"
    if "البرز" in s:
        return "البرز"
    if "سینا" in s:
        return "سینا"
    if "کوثر" in s:
        return "کوثر"
    if "sos" in s.lower():
        return "sos"
    if "پارسیان" in s:
        return "پارسیان"
    if "پاسارگاد" in s:
        return "پاسارگاد"
    if "صداوسیما" in s:
        return "صداوسیما"
    if "کارآفرین" in s or "کار افرین" in s:
        return "کارآفرین"
    if "ایران" in s:
        return "ایران"
    if "معلم" in s:
        return "معلم"
    if "آتیه سازان حافظ" in s:
        return "آتیه سازان حافظ"
    if "دانا" in s:
        return "دانا"
    return s

xlsx = Path(XLSX_PATH)
if not xlsx.exists():
    raise SystemExit(f"ERROR: Excel file not found: {XLSX_PATH}")

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = []
for r in ws.iter_rows(min_row=1, values_only=True):
    vals = [("" if v is None else str(v).strip()) for v in r[:7]]
    if len(vals) < 7:
        continue

    raw_name = vals[1]
    if not raw_name or raw_name == "نام بیمه":
        continue

    canonical = canon_insurer(raw_name)

    bucket = "unknown"
    weight = 60
    if vals[2] == "*":
        bucket, weight = "1m", 100
    elif vals[3] == "*":
        bucket, weight = "2m", 90
    elif vals[4] == "*":
        bucket, weight = "3m", 80
    elif vals[5] == "*":
        bucket, weight = "4m", 70
    elif vals[6] == "*":
        bucket, weight = "5m_plus", 55

    rows.append((raw_name, canonical, bucket, weight))

if not rows:
    raise SystemExit("ERROR: No rows parsed from Excel")

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

cur.executescript("""
DROP TABLE IF EXISTS insurance_reference_raw;
DROP TABLE IF EXISTS insurance_reference;

CREATE TABLE insurance_reference_raw (
    raw_name TEXT,
    canonical_name TEXT,
    speed_bucket TEXT,
    insurer_weight INTEGER
);

CREATE TABLE insurance_reference (
    canonical_name TEXT PRIMARY KEY,
    sample_raw_name TEXT,
    speed_bucket TEXT,
    insurer_weight INTEGER
);
""")

cur.executemany("""
INSERT INTO insurance_reference_raw (raw_name, canonical_name, speed_bucket, insurer_weight)
VALUES (?, ?, ?, ?)
""", rows)

cur.execute("""
INSERT INTO insurance_reference (canonical_name, sample_raw_name, speed_bucket, insurer_weight)
SELECT
    canonical_name,
    MIN(raw_name) AS sample_raw_name,
    MIN(speed_bucket) AS speed_bucket,
    MAX(insurer_weight) AS insurer_weight
FROM insurance_reference_raw
WHERE canonical_name IS NOT NULL
GROUP BY canonical_name
""")

conn.commit()

cur.execute("SELECT COUNT(*) FROM insurance_reference_raw")
raw_count = cur.fetchone()[0]
cur.execute("SELECT COUNT(*) FROM insurance_reference")
ref_count = cur.fetchone()[0]

print("insurance_reference loaded successfully")
print("raw rows:", raw_count)
print("canonical rows:", ref_count)

conn.close()
